import PySimpleGUI as sg
from bs4 import BeautifulSoup as bs
from requests import Session
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import winsound


sg.theme('DarkBlack')      # Add some color to the window

global s
s = Session()

# ----------- Create the 3 layouts this Window will display -----------
layout1=[[sg.Text('Porfavor ingrese sus credenciales del SGA', key='text')],
        [sg.Text('Username', size=(10, 1 ), key='username'), sg.InputText(size=(15, 1 ), key='usernameInput')],
        [sg.Text('Password', size=(10, 1), key='password'), sg.InputText(size=(15, 1 ), key='passwordInput')],
]

layout2 = [[sg.Text('Procesando...')],
           [sg.ProgressBar(90, orientation='h', size=(20, 20), key='progressbar')]]

# ----------- Create actual layout using Columns and a row of Buttons
layout = [[sg.Column(layout1, key='-COL1-'), sg.Column(layout2, visible=False, key='-COL2-')],
          [sg.Submit(key='submit'), sg.Cancel(key='cancel')]]

window = sg.Window('escobaUnemi', layout)

layout = 1  # The currently visible layout

def meh(username, password):
    window[f'-COL1-'].update(visible=False)  
    window[f'-COL2-'].update(visible=True)
    
    progress_bar = window['progressbar']
    
    #Url del login de pregradovirtual.unemi.edu.ec
    url = "https://pregradovirtual.unemi.edu.ec/login/index.php"

    #Automatización del login
    s.verify = False    #se conecta de manera HTTP en vez de HTTPS, esto es lo que crea todos los warnings
    site = s.get(url)
    bs_content = bs(site.content, "html.parser")
    token = bs_content.find("input", {"name":"logintoken"})["value"]
    login_data = {"username":str(username),"password":str(password), "logintoken":token}
    print(s.post(url, login_data))
    s.post(url, login_data)

    #Activacion y creación del excel 
    workbook = Workbook()
    sheet = workbook.active

    # Settings del excel generado
    sheet.column_dimensions["A"].width = 45
    sheet.column_dimensions["B"].width = 55
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 35
    sheet.column_dimensions["E"].width = 10

    #Proceso de recolleción de datos
    i = 1
        #Array de links de los sitios de calificaciones de cada materia 
    links = ['https://pregradovirtual.unemi.edu.ec/grade/report/user/index.php?id=319', 'https://pregradovirtual.unemi.edu.ec/grade/report/user/index.php?id=314', 'https://pregradovirtual.unemi.edu.ec/grade/report/user/index.php?id=315', 'https://pregradovirtual.unemi.edu.ec/grade/report/user/index.php?id=317', 'https://pregradovirtual.unemi.edu.ec/grade/report/user/index.php?id=318']

    #Definición de variables y herramintas para el proceso de recolección de datos
    arrayTareasLinks = []
    arrayTipo = []
    days = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    entregado = PatternFill(start_color= '7cc464', end_color= '7cc464', fill_type = "solid")

    for site in links:
        html = s.get(site)
        soup = bs( html.content, "lxml" )
        materia = soup.find("h3", class_="page-title")
        a = soup.find_all("a", class_="gradeitemheader")
        for minia in a:
            alts = minia.find_all("img", class_="itemicon")
            for alt in alts:
                progress_bar.UpdateBar(i + 1)
                #progress_bar.UpdateBar(i + 1)
                arrayTipo.append(alt["alt"])
                arrayTareasLinks.append(minia.attrs["href"])
                html2 = s.get(minia["href"])
                readable = bs(html2.content, "lxml")
                h2 = readable.find("h2")
                if alt["alt"] == "Foro":
                    fechas = readable.find("div", class_="alert-info")
                    date = fechas.text
                    
                    if any(day in date for day in days):
                        dateSplitted = date.split("es", 2)
                        datePrint = dateSplitted[2]
                    elif "límite" in date:
                        datePrint = "Cerrado"
                    nombreMateria = materia.text
                    sheet["A"+str(i)] = nombreMateria.split("-", 1)[0]
                    sheet["B"+str(i)] = h2.text
                    sheet["C"+str(i)] = alt["alt"]
                    sheet["D"+str(i)] = datePrint.strip()
                    sheet["E"+str(i)] = "ir a la tarea"
                    sheet["E"+str(i)].hyperlink = minia.attrs["href"]
                    i += 1
                if alt["alt"] == "Cuestionario":
                    pSearcher = readable.find("div", class_="quizinfo")                    
                    fechas = pSearcher.find_all("p")
                    spans = readable.find("span", class_="statedetails")
                    if spans is not None:
                        span = spans.text
                        status = span.split(":", 1)
                        if status[0] == "Enviado":
                            fecha = "Enviado, dont worry"
                    else:
                        for p in fechas:
                            d = p.text
                            if any(day in d for day in days):
                                date = d
                                fecha = date.split("el", 1)[1]
                    
                        
                    nombreMateria = materia.text
                    sheet["A"+str(i)] = nombreMateria.split("-", 1)[0]
                    sheet["B"+str(i)] = h2.text
                    sheet["C"+str(i)] = alt["alt"]
                    sheet["D"+str(i)] = fecha
                    sheet["E"+str(i)] = "ir a la tarea"
                    sheet["E"+str(i)].hyperlink = minia.attrs["href"]
                    if fecha == "Enviado, dont worry":
                        sheet["D"+str(i)].fill = entregado

                    i += 1
                    
                if alt["alt"] == "Tarea":
                    table = readable.find_all("td", attrs={"class": "c1"})
                    if table[1].text == "Enviado para calificar" or table[0].text == "Enviado para calificar":
                        nombreMateria = materia.text
                        sheet["A"+str(i)] = nombreMateria.split("-", 1)[0]
                        sheet["B"+str(i)] = h2.text
                        sheet["C"+str(i)] = alt["alt"]
                        sheet["D"+str(i)] = "Entregado, dont worry"
                        sheet["D"+str(i)].fill = entregado
                        sheet["E"+str(i)] = "ir a la tarea"
                        sheet["E"+str(i)].hyperlink = minia.attrs["href"]
                        i += 1
                    else:
                        nombreMateria = materia.text
                        sheet["A"+str(i)] = nombreMateria.split("-", 1)[0]
                        sheet["B"+str(i)] = h2.text
                        sheet["C"+str(i)] = alt["alt"]
                        sheet["D"+str(i)] = table[2].text
                        sheet["E"+str(i)] = "ir a la tarea"
                        sheet["E"+str(i)].hyperlink = minia.attrs["href"]
                        i += 1
        
    #Verfica si la celda a1 esta vacia y si es asi, da un error de login
    a1 = sheet["A1"]
    while a1.value == None:
        window[f'-COL1-'].update(visible=True)  
        window[f'-COL2-'].update(visible=False)
        winsound.PlaySound('*', winsound.SND_ALIAS)
        window.Element('text').Update('Hubo un error con las credenciales, porfavor intente de nuevo.')
        print('error')
        while True:
            event,values = window.read()
            if event in (None, 'cancel'):
                quit()
            if event == 'submit':
                
                username = values['usernameInput']
                password = values['passwordInput']
                
                meh(username, password)
    else:
        window[f'-COL1-'].update(visible=True)  
        window[f'-COL2-'].update(visible=False)
        window.Element('text').Update('El archivo horarioUnemi.xlsx ha sido generado exitosamente.')
        window.Element('username').Update(visible=False)
        window.Element('password').Update(visible=False)
        window.Element('usernameInput').Update(disabled=True, visible=False)
        window.Element('passwordInput').Update(disabled=True, visible=False)
        window.Element('submit').Update(disabled=True, visible=False)
        workbook.save(filename="horarioUnemi.xlsx")




while True:
    event,values = window.read()
    if event in (None, 'cancel'):
        quit()
    if event == 'submit':
        username = values['usernameInput']
        password = values['passwordInput']
        
        meh(username, password)